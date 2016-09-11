import openpyxl
from tools import *
from globals import regions_global


# This loads all the cases and regions into a freshly created database
def main():
    load_cases_to_database(get_cases_ascii('reise.cases'))
    load_regions_to_database(regions_global)


# Extracting all the cases from the ascii-file
def get_cases_ascii(file_name):
    f = open(file_name, 'r')
    cases = {}
    for line in f.readlines():
        line = line.strip()
        columns = line.split()
        line_length = len(columns)
        if line_length >= 1:
            name = columns[0]
            if name == 'defcase':
                cases[columns[1]] = []
                defcase = columns[1]
            elif name == 'case':
                cases[defcase].append(columns[1])
            elif name == 'JourneyCode:':
                cases[defcase].append(columns[1])
            elif name == 'HolidayType:':
                cases[defcase].append(columns[1].replace(',', ''))
            elif name == 'Price:':
                cases[defcase].append(columns[1])
            elif name == 'NumberOfPersons:':
                cases[defcase].append(columns[1])
            elif name == 'Region:':
                cases[defcase].append(columns[1].replace(',', ''))
            elif name == 'Transportation:':
                cases[defcase].append(columns[1].replace(',', ''))
            elif name == 'Duration:':
                cases[defcase].append(columns[1])
            elif name == 'Season:':
                cases[defcase].append(columns[1].replace(',', ''))
            elif name == 'Accommodation:':
                cases[defcase].append(columns[1].replace(',', ''))
            elif name == 'Hotel:':
                number_of_columns = 1
                hotel_name = ''
                while number_of_columns < line_length:
                    hotel_name += columns[number_of_columns] + ' '
                    number_of_columns += 1
                cases[defcase].append(hotel_name.replace('"', '')[:-1])
            else:
                pass
    f.close()
    return cases


# Extracting all the cases from the excel-file
def get_cases_excel(file_name):
    wb = openpyxl.load_workbook(file_name)
    sheet_names = wb.get_sheet_names()
    sheet_name = sheet_names[0]
    sheet = wb.get_sheet_by_name(sheet_name)
    case_column = 3
    case_row = 1
    cases = {}
    is_case = True
    while is_case:
        case = [None] * 11
        header_cell = sheet.cell(row=case_row, column=case_column-2).value
        if header_cell is None or header_cell != 'defcase':
            is_case = False
        else:
            for next_cell in range(2, 13):
                case[next_cell-2] = sheet.cell(row=case_row + next_cell, column=3).value
            case[1] = str(case[1])
            case[2] = case[2].replace(',', '')
            case[3] = str(case[3])
            case[4] = str(case[4])
            case[5] = case[5].replace(',', '')
            case[6] = case[6].replace(',', '')
            case[7] = str(case[7])
            case[8] = case[8].replace(',', '')
            case[9] = case[9].replace(',', '')
            cases[case[1]] = case
        case_row += 16
    return cases


if __name__ == '__main__':
    main()

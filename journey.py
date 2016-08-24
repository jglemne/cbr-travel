#!/usr/local/bin/python3.5

# import sys
# import platform
import openpyxl
import operator
# import time
from geopy.distance import great_circle
import tkinter as tk
import tkinter.font as tkFont
import tkinter.ttk as ttk
from globals import fields_global, drop_downs_global
from tools import *
from features import *


def main():
    target_case = TargetCase()
    instance_cases(retrieve_cases(), target_case)
    Interface(target_case).mainloop()


class MultiColumnListbox:

    def __init__(self, nr_of_results):
        self.tree = None
        self._setup_widgets()
        self.build_tree(nr_of_results)

    def _setup_widgets(self):
        container = ttk.Frame()
        container.pack(fill='both', expand=True, padx=(10, 10), pady=(10, 10))
        # create a treeview with dual scrollbars
        self.tree = ttk.Treeview(columns=JourneyCase.list, show="headings")
        vsb = ttk.Scrollbar(orient="vertical",
                            command=self.tree.yview)
        hsb = ttk.Scrollbar(orient="horizontal",
                            command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set,
                            xscrollcommand=hsb.set)
        self.tree.grid(column=0, row=0, sticky='nsew', in_=container)
        vsb.grid(column=1, row=0, sticky='ns', in_=container)
        hsb.grid(column=0, row=1, sticky='ew', in_=container)
        container.grid_columnconfigure(0, weight=1)
        container.grid_rowconfigure(0, weight=1)

    def build_tree(self, nr_of_results):
        for col in JourneyCase.list:
            self.tree.heading(col, text=col, command=lambda c=col: sortby(self.tree, c, 0))
            # adjust the column's width to the header string
            self.tree.column(col, width=tkFont.Font().measure(col))
        for item in JourneyCase.similarities()[0:nr_of_results]:
            self.tree.insert('', 'end', values=item[0].case_tuples)
            # adjust column's width if necessary to fit each value
            # for ix, val in enumerate(item):
            #     col_w = tkFont.Font().measure(val)
            #     if self.tree.column(JourneyCase.list[ix], width=None) < col_w:
            #         self.tree.column(JourneyCase.list[ix], width=col_w)


class Field:
    fields = {}

    @classmethod
    def create(cls, master, label_text, row, col):
        instance = Field(master, label_text, row, col)
        cls.fields[label_text] = instance
        return instance

    def __init__(self, master, label_text, row, col):
        self.label = tk.Label(master, text=label_text)
        self.entry = tk.Entry(master)
        self.row = row
        self.col = col

    def make_grid(self):
        self.label.grid(row=self.row, sticky=tk.E)
        self.entry.grid(row=self.row, column=self.col)

    def get_input(self):
        return self.entry.get()

    input = property(get_input)


class DropDown:
    drop_downs = {}

    @classmethod
    def create(cls, master, label_text, row, col):
        instance = DropDown(master, label_text, row, col)
        cls.drop_downs[label_text] = instance
        return instance

    def __init__(self, master, label_text, row, col):
        self.var = tk.StringVar(master)
        self.label = tk.Label(master, text=label_text)
        self.entry = tk.OptionMenu(master, self.var, *drop_downs_global[label_text])
        self.row = row
        self.col = col

    def make_grid(self):
        self.label.grid(row=self.row, sticky=tk.E)
        self.entry.grid(row=self.row, column=self.col, sticky=tk.EW)

    def get_input(self):
        return self.var.get()

    input = property(get_input)


class Interface(tk.Tk):
    field_row = 0
    field_column = 1
    entries = {}
    nr_of_results = 100

    def __init__(self, target_case):
        tk.Tk.__init__(self)
        self.title('CBR Travel Case')
        self.menu_frame = tk.Frame(self)
        self.menu_frame.pack(side=tk.TOP)
        self.menu = tk.Menu(self.menu_frame)
        self.file_menu = tk.Menu(self.menu, tearoff=0)
        self.file_menu.add_command(label="Add new case")
        self.file_menu.add_command(label="Edit case")
        self.file_menu.add_command(label="Delete case")
        self.file_menu.add_command(label="Edit weights")
        self.file_menu.add_command(label="Close")
        self.file_menu.add_separator()
        self.file_menu.add_command(label="Quit program", command=self.quit)
        self.menu.add_cascade(label="Case Base", menu=self.file_menu)
        self.config(menu=self.menu)
        welcome = "Welcome! \n" \
            "This CBR-system is designed to give you journey suggestions based " \
            "on your preferences. When you edit your preferences here below and then press" \
            "'Get best matches', the system will compare your preferences with a given " \
            "number of cases and bring you the best matches calculated in percent. "
        self.msg = ttk.Label(wraplength="8i", justify="center", anchor="n", padding=(10, 10, 10, 10), text=welcome)
        self.msg.pack(fill='x')
        self.field_frame = tk.Frame(self)
        self.field_frame.pack(side=tk.TOP, pady=(10, 0))
        for field in fields_global:
            self.entries[field] = Field.create(self.field_frame, field, self.field_row, self.field_column)
            self.entries[field].make_grid()
            self.field_row += 1
        for drop_down in drop_downs_global:
            self.entries[drop_down] = DropDown.create(self.field_frame, drop_down, self.field_row, self.field_column)
            self.entries[drop_down].make_grid()
            self.field_row += 1
        self.button = tk.Button(self.field_frame, text="Get best matches", command=self.on_button)
        self.button.grid(columnspan=2, pady=(10, 0))
        self.list = MultiColumnListbox(self.nr_of_results)
        self.target_case = target_case

    def on_button(self):
        self.list.tree.delete(*self.list.tree.get_children())
        for entry in self.entries:
            set_target_case_feature(entry, self.entries[entry].input, self.target_case)
        self.list.build_tree(self.nr_of_results)


def donothing(root):
    filewin = tk.Toplevel(root)
    # button = tk.Button(filewin, text="Do nothing button")
    # button.pack()


class TargetCase:

    def __init__(
            self, case=None, journey_code=None, holiday_type=None,
            price=None, number_of_persons=None, region=None, transportation=None,
            duration=None, season=None, accommodation=None, hotel=None):
        self.accommodation = Accommodation(accommodation)
        self.case = case
        self.duration = Duration(duration)
        self.holiday_type = HolidayType(holiday_type)
        self.hotel = Hotel(hotel)
        self.journey_code = JourneyCode(journey_code)
        self.number_of_persons = NumberOfPersons(number_of_persons)
        self.price = Price(price, number_of_persons)
        self.region = Region(region)
        self.season = Season(season)
        self.transportation = Transportation(transportation)

    def get_case_list(self):
        return [
            self.case,
            self.journey_code.number,
            self.holiday_type.name,
            self.price.total,
            self.number_of_persons.total,
            self.region.name,
            self.transportation.name,
            self.duration.days,
            self.season.month,
            self.accommodation.name,
            self.hotel.name
        ]

    def get_accommodation(self):
        return self.accommodation.name

    def set_accommodation(self, accommodation):
        self.accommodation = Accommodation(accommodation)

    def del_accommodation(self):
        del self.accommodation
        self.accommodation = Accommodation()

    def get_case(self):
        return self.case

    def set_case(self, case):
        self.case = case

    def del_case(self):
        del self.case

    def get_duration(self):
        return self.duration.days

    def set_duration(self, duration):
        self.duration = Duration(duration)

    def del_duration(self):
        del self.duration
        self.duration = Duration()

    def get_holiday_type(self):
        return self.holiday_type.name

    def set_holiday_type(self, holiday_type):
        self.holiday_type = HolidayType(holiday_type)

    def del_holiday_type(self):
        del self.holiday_type
        self.holiday_type = HolidayType()

    def get_hotel(self):
        return self.hotel.name

    def set_hotel(self, hotel):
        self.hotel = Hotel(hotel)

    def del_hotel(self):
        del self.hotel
        self.hotel = Hotel()

    def get_journey_code(self):
        return self.journey_code.number

    def set_journey_code(self, journey_code):
        self.journey_code = JourneyCode(journey_code)

    def del_journey_code(self):
        del self.journey_code
        self.journey_code = JourneyCode()

    def get_number_of_persons(self):
        return self.number_of_persons.total

    def set_number_of_persons(self, number_of_persons):
        self.number_of_persons = NumberOfPersons(number_of_persons)

    def del_number_of_persons(self):
        del self.number_of_persons
        self.number_of_persons = NumberOfPersons()

    def get_price(self):
        return self.price.total

    def set_price(self, price):
        self.price = Price(price)

    def del_price(self):
        del self.price
        self.price = Price()

    def get_region(self):
        return self.region.name

    def set_region(self, region):
        self.region = Region(region)

    def del_region(self):
        del self.region
        self.region = Region()

    def get_season(self):
        return self.season.month

    def set_season(self, season):
        self.season = Season(season)

    def del_season(self):
        del self.season
        self.season = Season()

    def get_transportation(self):
        return self.transportation.name

    def set_transportation(self, transportation):
        self.transportation = Transportation(transportation)

    def del_transportation(self):
        del self.transportation
        self.transportation = Transportation()

    list = property(get_case_list)
    a = property(get_accommodation, set_accommodation, del_accommodation)
    c = property(get_case, set_case, del_case)
    d = property(get_duration, set_duration, del_duration)
    ht = property(get_holiday_type, set_holiday_type, del_holiday_type)
    h = property(get_hotel, set_hotel, del_hotel)
    jc = property(get_journey_code, set_journey_code, del_journey_code)
    nop = property(get_number_of_persons, set_number_of_persons, del_number_of_persons)
    p = property(get_price, set_price, del_price)
    r = property(get_region, set_region, del_region)
    s = property(get_season, set_season, del_season)
    t = property(get_transportation, set_transportation, del_transportation)


class JourneyCase:
    cases = {}
    max_code = 0
    holiday_types = {}
    holiday_list = []
    prices = []
    persons_per_case = []
    price_range = [239, 8007]
    durations = []
    accommodations = {}
    hotels = {}
    regions = {}
    seasons = {}
    transportations = {}
    list = [
        'Similarity [%]',
        'Case [index]',
        'Holiday type',
        'Price [NZD]',
        'Persons [#]',
        'Region in the world',
        'Transport',
        'Duration [days]',
        'Season [month]',
        'Accommodation type',
        'Hotel name'
        ]

    @classmethod
    def create(
            cls, case, journey_code, holiday_type,
            price, number_of_persons, region, transportation,
            duration, season, accommodation, hotel, target_case):
        instance = JourneyCase(
            case, journey_code, holiday_type,
            price, number_of_persons, region, transportation,
            duration, season, accommodation, hotel, target_case)
        cls.cases[instance] = instance
        if cls.max_code < int(instance.journey_code.number):
            cls.max_code = int(instance.journey_code.number)
        if holiday_type not in cls.holiday_types:
            cls.holiday_types[holiday_type] = holiday_type
        cls.persons_per_case.append(instance.number_of_persons.total)
        if accommodation not in cls.accommodations:
            cls.accommodations[accommodation] = accommodation
        if instance.duration.days not in cls.durations:
            cls.durations.append(instance.duration.days)
        if hotel not in cls.hotels:
            cls.hotels[hotel] = hotel
        if region not in cls.regions:
            cls.regions[region] = region
        if season not in cls.seasons:
            cls.seasons[season] = season
        if transportation not in cls.transportations:
            cls.transportations[transportation] = transportation
        cls.prices.append(price)
        return instance

    @classmethod
    def similarities(cls):
        similarities = {}
        for instance in cls.cases:
            similarities[instance] = instance.similarity()
        return sorted(similarities.items(), key=operator.itemgetter(1), reverse=True)

    def __init__(
            self, case=None, journey_code=None, holiday_type=None,
            price=None, number_of_persons=None, region=None, transportation=None,
            duration=None, season=None, accommodation=None, hotel=None, target_case=TargetCase()):
        self.accommodation = Accommodation(accommodation)
        self.case = case
        self.duration = Duration(duration)
        self.holiday_type = HolidayType(holiday_type)
        self.hotel = Hotel(hotel)
        self.journey_code = JourneyCode(journey_code)
        self.number_of_persons = NumberOfPersons(number_of_persons)
        self.price = Price(price, number_of_persons)
        self.region = Region(region)
        self.season = Season(season)
        self.transportation = Transportation(transportation)
        self.target_case = target_case
        self.similarity()

    def get_case_features(self):
        return {
            'Case': self.journey_code.number,
            'Holiday type': self.holiday_type.name,
            'Price': self.price.total,
            'Number of persons': self.number_of_persons.total,
            'Region': self.region.name.title(),
            'Transportation': self.transportation.name,
            'Duration': self.duration.days,
            'Season': self.season.month,
            'Accommodation': self.accommodation.name,
            'Hotel': self.hotel.name
        }

    def get_case_tuple(self):
        return (
            "{0:.2f}".format(round(self.similarity()*100, 2)),
            self.journey_code.number,
            self.holiday_type.name,
            self.price.total,
            self.number_of_persons.total,
            self.region.name.title(),
            self.transportation.name,
            self.duration.days,
            self.season.month,
            self.accommodation.name,
            self.hotel.name
        )

    features = property(get_case_features)
    case_tuples = property(get_case_tuple)

    def similarity(self):
        sim_int = 0
        total_weight = 0
        # Accommodation
        if self.target_case.accommodation.name is not None:
            weight = self.accommodation.weight
            sim_int += self.accommodation_sim() * weight
            total_weight += weight
        else:
            sim_int += 1
            total_weight += 1
        # Duration
        if self.target_case.duration.days is not None:
            weight = self.duration.weight
            sim_int += self.duration_sim() * weight
            total_weight += weight
        else:
            sim_int += 1
            total_weight += 1
        # Holiday type
        if self.target_case.holiday_type.name is not None:
            weight = self.holiday_type.weight
            sim_int += self.holiday_type_sim() * weight
            total_weight += weight
        else:
            sim_int += 1
            total_weight += 1
        # Hotel
        if self.target_case.hotel.name is not None:
            weight = self.hotel.weight
            sim_int += self.hotel_sim() * weight
            total_weight += weight
        else:
            sim_int += 1
            total_weight += 1
        # Journey code
        if self.target_case.journey_code.number is not None:
            weight = self.journey_code.weight
            sim_int += self.journey_code_sim() * weight
            total_weight += weight
        else:
            sim_int += 1
            total_weight += 1
        # Number of persons
        if self.target_case.number_of_persons.total is not None:
            weight = self.number_of_persons.weight
            sim_int += self.number_of_persons_sim() * weight
            total_weight += weight
        else:
            sim_int += 1
            total_weight += 1
        # Price
        if self.target_case.price.total is not None:
            weight = self.price.weight
            sim_int += self.price_sim() * weight
            total_weight += weight
        else:
            sim_int += 1
            total_weight += 1
        # Region
        if self.target_case.region.name is not None:
            weight = self.region.weight
            sim_int += self.region_sim() * weight
            total_weight += weight
        else:
            sim_int += 1
            total_weight += 1
        # Season
        if self.target_case.season.name is not None:
            weight = self.season.weight
            sim_int += self.season_sim() * weight
            total_weight += weight
        else:
            sim_int += 1
            total_weight += 1
        # Transportation
        if self.target_case.transportation.similarity is not None:
            weight = self.transportation.weight
            sim_int += self.transportation_sim() * weight
            total_weight += weight
        else:
            sim_int += 1
            total_weight += 1
        return sim_int/total_weight

    def accommodation_sim(self):
        if self.accommodation.index >= self.target_case.accommodation.index:
            return 1
        else:
            return self.accommodation.index/self.target_case.accommodation.index

    def duration_sim(self):
        if self.duration.days == self.target_case.duration.days:
            return 1
        elif abs(self.duration.days - self.target_case.duration.days) <= 4:
            return (5 - abs(self.duration.days - self.target_case.duration.days))/5
        else:
            return 0

    def holiday_type_sim(self):
        if self.holiday_type.group == self.target_case.holiday_type.group:
            return 1
        elif self.holiday_type.group[-3:] == self.target_case.holiday_type.group[-3:]:
            if self.target_case.holiday_type.group[-3:] == '100':
                return 0.6
            else:
                return 0.5
        else:
            return 0.3

    def hotel_sim(self):
        if self.hotel.name == self.target_case.hotel.name:
            return 1
        else:
            return 0

    def journey_code_sim(self):
        if self.journey_code.number == self.target_case.journey_code.number:
            return 1
        else:
            return 0

    def number_of_persons_sim(self):
        if self.number_of_persons.total == self.target_case.number_of_persons.total:
            return 1
        elif abs(self.number_of_persons.total - self.target_case.number_of_persons.total) < 2:
            return 0.5
        else:
            return 0

    def price_sim(self):
        if (self.price.total <= self.target_case.price.total) | (self.target_case.price.total > self.price_range[1]):
            return 1
        elif self.target_case.price.total < self.price_range[0]:
            return 0
        else:
            return 1 - ((self.price.total - self.target_case.price.total) / (self.price_range[1] - self.price_range[0]))

    def region_sim(self):
        if self.region.name == self.target_case.region.name:
            return 1
        elif self.target_case.region.coordinates is not None:
            target = (self.target_case.region.coordinates['Lat'], self.target_case.region.coordinates['Long'])
            source = (self.region.coordinates['Lat'], self.region.coordinates['Long'])
            distance = great_circle(target, source).kilometers
            if distance > self.region.distance:
                return 0
            else:
                lat_distance = great_circle((target[0], 0), (source[0], 0)).kilometers
                latitude_sim = (self.region.distance-lat_distance)/self.region.distance
                distance_sim = (self.region.distance-distance)/self.region.distance
                return (distance_sim + (2 * latitude_sim))/3
        else:
            return 0

    def season_sim(self):
        if self.season.month == self.target_case.season.month:
            return 1
        elif self.season.name[1] == self.target_case.season.name[1]:
            return 0.5
        elif self.season.name[0] == self.target_case.season.name[0]:
            return 0.2
        else:
            return 0

    def transportation_sim(self):
        return self.target_case.transportation.similarity[self.transportation.similarity.index(1.0)]


def instance_cases(cases, target_case):
    for case in cases:
        case = list(case)
        JourneyCase.create(
            case[1], case[0], case[3], int(case[4]),
            int(case[5]), case[6], case[7], int(case[8]),
            case[9], case[10], case[11], target_case)


def load_cases_excel(file_name, target_case):
    wb = openpyxl.load_workbook(file_name)
    sheet_names = wb.get_sheet_names()
    sheet_name = sheet_names[0]
    sheet = wb.get_sheet_by_name(sheet_name)
    case_column = 3
    case_row = 1
    is_case = True
    while is_case:
        case = [None] * 11
        header_cell = sheet.cell(row=case_row, column=case_column-2).value
        if header_cell is None or header_cell != 'defcase':
            is_case = False
        else:
            for next_cell in range(2, 13):
                case[next_cell-2] = sheet.cell(row=case_row + next_cell, column=3).value
            JourneyCase.create(
                case[0], case[1], case[2].replace(',', ''), case[3],
                case[4], case[5].replace(',', ''), case[6].replace(',', ''), case[7],
                case[8].replace(',', ''), case[9].replace(',', ''), case[10], target_case)
        case_row += 16


def set_target_case_feature(feature_name, feature_entry, target_case):
    if feature_name == 'Accommodation':
        if feature_entry in JourneyCase.accommodations:
            target_case.a = feature_entry
        else:
            del target_case.a
    elif feature_name == 'Duration':
        if feature_entry != '':
            try:
                target_case.d = int(feature_entry)
            except ValueError:
                del target_case.d
        else:
            del target_case.d
    elif feature_name == 'Holiday type':
        if feature_entry in JourneyCase.holiday_types:
            target_case.ht = feature_entry
        else:
            del target_case.ht
    elif feature_name == 'Hotel name':
        if feature_entry in JourneyCase.hotels:
            target_case.h = feature_entry
        else:
            del target_case.h
    elif feature_name == 'Number of persons':
        if feature_entry != '':
            try:
                target_case.nop = int(feature_entry)
            except ValueError:
                del target_case.nop
        else:
            del target_case.nop
    elif feature_name == 'Price':
        if feature_entry != '':
            try:
                target_case.p = int(feature_entry)
            except ValueError:
                del target_case.p
        else:
            del target_case.p
    elif feature_name == 'Region':
        if feature_entry != '':
            target_case.r = feature_entry
        else:
            del target_case.r
    elif feature_name == 'Season':
        if feature_entry in JourneyCase.seasons:
            target_case.s = feature_entry
        else:
            del target_case.s
    elif feature_name == 'Transportation':
        if feature_entry in JourneyCase.transportations:
            target_case.t = feature_entry
        else:
            del target_case.t
    else:
        pass


# Standard boilerplate to call the main() function to begin
# the program.
if __name__ == '__main__':
    main()

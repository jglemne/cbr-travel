import openpyxl


def get_cases_ascii(file_name):
    f = open(file_name, 'r')
    cases = {}
    for line in f.readlines():
        line = line.strip()
        columns = line.split()
        line_length = len(columns)
        if line_length >= 1:
            name = columns[0]
            if name == 'defcase':
                cases[columns[1]] = []
                defcase = columns[1]
            elif name == 'case':
                cases[defcase].append(columns[1])
            elif name == 'JourneyCode:':
                cases[defcase].append(columns[1])
            elif name == 'HolidayType:':
                cases[defcase].append(columns[1].replace(',', ''))
            elif name == 'Price:':
                cases[defcase].append(columns[1])
            elif name == 'NumberOfPersons:':
                cases[defcase].append(columns[1])
            elif name == 'Region:':
                cases[defcase].append(columns[1].replace(',', ''))
            elif name == 'Transportation:':
                cases[defcase].append(columns[1].replace(',', ''))
            elif name == 'Duration:':
                cases[defcase].append(columns[1])
            elif name == 'Season:':
                cases[defcase].append(columns[1].replace(',', ''))
            elif name == 'Accommodation:':
                cases[defcase].append(columns[1].replace(',', ''))
            elif name == 'Hotel:':
                number_of_columns = 1
                hotel_name = ''
                while number_of_columns < line_length:
                    hotel_name += columns[number_of_columns] + ' '
                    number_of_columns += 1
                cases[defcase].append(hotel_name.replace('"', '')[:-1])
            else:
                pass
    f.close()
    return cases


def count_cases(sheet):
    is_case = True
    cell_letter = 'A'
    cell_number = '1'
    number_of_cases = 0
    while is_case:
        if sheet[cell_letter + cell_number].value is None or sheet[cell_letter + cell_number].value != 'defcase':
            is_case = False
        else:
            number_of_cases += 1
            cell_int = int(cell_number)
            cell_int += 16
            cell_number = str(cell_int)
    return number_of_cases


def get_cases_price_interval(lowest_price, highest_price, my_dict):
    indices = []
    for price in range(lowest_price, highest_price):
        empty_dict = False
        while not empty_dict:
            if price in my_dict.values():
                case = list(my_dict.keys())[list(my_dict.values()).index(price)]
                indices.append(case)
                my_dict = remove_key(my_dict, case)
            else:
                empty_dict = True
    return indices


def remove_key(d, key):
    r = dict(d)
    del r[key]
    return r


def hamming_distance(s1, s2):
    if len(s1) != len(s2):
        raise ValueError("Undefined for sequences of unequal length")
    return sum(el1 != el2 for el1, el2 in zip(s1, s2))


def load_cases_excel(file_name, target_case):
    wb = openpyxl.load_workbook(file_name)
    sheet_names = wb.get_sheet_names()
    sheet_name = sheet_names[0]
    sheet = wb.get_sheet_by_name(sheet_name)
    case_column = 3
    case_row = 1
    is_case = True
    while is_case:
        case = [None] * 11
        header_cell = sheet.cell(row=case_row, column=case_column-2).value
        if header_cell is None or header_cell != 'defcase':
            is_case = False
        else:
            for next_cell in range(2, 13):
                case[next_cell-2] = sheet.cell(row=case_row + next_cell, column=3).value
            JourneyCase.create(
                case[0], case[1], case[2].replace(',', ''), case[3],
                case[4], case[5].replace(',', ''), case[6].replace(',', ''), case[7],
                case[8].replace(',', ''), case[9].replace(',', ''), case[10], target_case)
        case_row += 16